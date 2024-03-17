import CEC2005.functions as cec2005
import time

from matplotlib import pyplot as plt
from mealpy import GA
from openpyxl import load_workbook

# 定义目标函数
fitness_function = cec2005.fun1

# 维度
dim = 30

# 搜索空间范围
bounds = 10

# 最大迭代次数
epoch = 2000

# 种群数量
pop_size = 50

# Excel 数据行起始位置
row = 1

# Excel 数据列起始位置
col = 'A'

# 循环次数
count = 15

problem = {
    "fit_func": fitness_function,
    "lb": [-bounds, ] * dim,
    "ub": [bounds, ] * dim,
    "minmax": "min",
    "log_to": None
}

for _ in range(count):
    '''GA'''
    time_start = time.time()
    ga_model = GA.BaseGA(epoch, pop_size, pc=0.9, pm=0.05)
    ga_best_x, ga_best_f = ga_model.solve(problem)
    time_end = time.time()
    ga_cost = time_end - time_start

    ''' 输出结果 '''
    # 绘制适应度曲线
    function_name = fitness_function.__name__
    filename = f"{'CEC2005 ' + function_name + '_GA_' + str(row)}.jpg"
    plt.figure(figsize=(8, 6), dpi=300)

    plt.plot(ga_model.history.list_global_best_fit, 'r', linewidth=2, label='GA')

    plt.title(f'CEC2005 ' + function_name + ' Convergence curve: ', fontsize=15)
    plt.xlabel('Iteration')  # 设置x轴标签
    plt.ylabel('Fitness')  # 设置y轴标签

    plt.grid()  # 显示网格
    plt.legend()  # 显示图例
    # plt.savefig('Data/' + filename)  # 保存图像
    # plt.show()

    # 打开现有的工作簿
    workbook = load_workbook('Data/testdata.xlsx')

    # 选择指定的工作表
    if function_name in workbook.sheetnames:
        sheet = workbook[function_name]
    else:
        sheet = workbook.create_sheet(function_name)

    # 将数据写入指定位置
    sheet[col + str(row)] = ga_best_f
    sheet[chr(ord(col) + 1) + str(row)] = ga_cost

    # 保存工作簿
    workbook.save('Data/testdata.xlsx')
    print(str(row) + '/' + str(count))

    row += 1

# 播放提示音
# os.system('afplay /Users/lovir/Music/三全音.aif')

print(function_name + " is done!")
